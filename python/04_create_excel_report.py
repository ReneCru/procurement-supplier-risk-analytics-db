"""
Create an Excel report from dashboard-ready CSV exports.

Output:
reports/procurement_supplier_analytics_report.xlsx
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment


PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_DIR = PROJECT_ROOT / "data" / "dashboard_exports"
REPORT_DIR = PROJECT_ROOT / "reports"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_FILE = REPORT_DIR / "procurement_supplier_analytics_report.xlsx"


FILES = {
    "Supplier Risk": "supplier_risk_score.csv",
    "Supplier Spend": "supplier_spend_summary.csv",
    "Spend Category": "spend_by_category.csv",
    "Monthly Trend": "monthly_purchasing_trend.csv",
    "Late Delivery": "late_delivery_performance.csv",
    "Open PO Aging": "open_po_aging.csv",
    "Compliance": "supplier_compliance_status.csv",
}


def autosize_columns(ws):
    """Autosize worksheet columns."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 45)


def format_header(ws):
    """Apply basic header formatting."""
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def create_summary_sheet(writer):
    """Create executive summary sheet."""
    summary_data = {
        "Metric": [
            "Total Suppliers",
            "Total Buyers",
            "Total Categories",
            "Total Purchase Orders",
            "Total PO Lines",
            "Total Deliveries",
            "Total Invoices",
            "Total Compliance Documents",
        ],
        "Value": [
            165,
            24,
            9,
            500,
            1982,
            1982,
            500,
            990,
        ],
    }

    df = pd.DataFrame(summary_data)
    df.to_excel(writer, sheet_name="Executive Summary", index=False)


def main():
    """Create Excel workbook."""
    print("Creating Excel procurement analytics report...")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        create_summary_sheet(writer)

        for sheet_name, file_name in FILES.items():
            file_path = EXPORT_DIR / file_name

            if not file_path.exists():
                raise FileNotFoundError(f"Missing dashboard export: {file_path}")

            df = pd.read_csv(file_path)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(OUTPUT_FILE)

    for ws in wb.worksheets:
        format_header(ws)

    # Chart 1: Spend by Category
    ws = wb["Spend Category"]
    if ws.max_row > 1:
        chart = BarChart()
        chart.title = "Spend by Category"
        chart.y_axis.title = "Total Spend"
        chart.x_axis.title = "Category"

        data = Reference(ws, min_col=7, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20

        ws.add_chart(chart, "J2")

    # Chart 2: Monthly Purchasing Trend
    ws = wb["Monthly Trend"]
    if ws.max_row > 1:
        chart = LineChart()
        chart.title = "Monthly Purchasing Trend"
        chart.y_axis.title = "Total Spend"
        chart.x_axis.title = "Month"

        data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20

        ws.add_chart(chart, "H2")

    # Chart 3: Open PO Aging
    ws = wb["Open PO Aging"]
    if ws.max_row > 1:
        aging_summary = (
            pd.read_csv(EXPORT_DIR / "open_po_aging.csv")
            .groupby("aging_bucket", as_index=False)["total_amount"]
            .sum()
            .rename(columns={"total_amount": "open_po_value"})
        )

        start_row = ws.max_row + 3
        ws.cell(row=start_row, column=1, value="Aging Bucket")
        ws.cell(row=start_row, column=2, value="Open PO Value")

        for index, row in aging_summary.iterrows():
            ws.cell(row=start_row + index + 1, column=1, value=row["aging_bucket"])
            ws.cell(row=start_row + index + 1, column=2, value=float(row["open_po_value"]))

        chart = BarChart()
        chart.title = "Open PO Aging"
        chart.y_axis.title = "Open PO Value"
        chart.x_axis.title = "Aging Bucket"

        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(aging_summary))
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(aging_summary))

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20

        ws.add_chart(chart, "J2")

    wb.save(OUTPUT_FILE)

    print(f"Excel report created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
